"""
ETL: Parse all 4 Excel tables and load into SQLite database.
Usage: python etl_sync_all.py
"""
import sqlite3
import re
import os
import json
from datetime import datetime
from openpyxl import load_workbook

# ============================================================
# CONFIG — 优先从 .env 读取，没有则提示用户配置
# ============================================================
DB_PATH = os.environ.get('DB_PATH', os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'ur_esports.db'))
DATA_DIR = os.environ.get('DATA_DIR', '').strip()

if not DATA_DIR:
    # 尝试从后端 .env 文件读取
    env_file = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '.env')
    if os.path.exists(env_file):
        with open(env_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line.startswith('DATA_DIR='):
                    DATA_DIR = line.split('=', 1)[1].strip().strip('"').strip("'")
                    break

if not DATA_DIR:
    print("__JSON__START__")
    print(json.dumps({"error": "DATA_DIR 未配置，请在前端「赛训导入」页面设置数据文件夹路径"}, ensure_ascii=False))
    print("__JSON__END__")
    exit(0)

TRAINING_LOG = os.path.join(DATA_DIR, "UR_CS2_训练日志.xlsx")
BRIEFING = os.path.join(DATA_DIR, "UR_CS2_每日简报.xlsx")
TACTICS = os.path.join(DATA_DIR, "UR_CS2_战术总表.xlsx")
MATCH_DATA = os.path.join(DATA_DIR, "CS2_Match_Data_May2026.xlsx")

# Player ID mapping loaded dynamically from database at runtime
PLAYER_MAP = {}

MAP_ALIASES = {
    'overpass': 'Overpass', 'op': 'Overpass',
    'ancient': 'Ancient', 'anc': 'Ancient',
    'd2': 'Dust2', 'd3': 'Dust2', 'd4': 'Dust2', 'd5': 'Dust2',
    'd6': 'Dust2', 'd7': 'Dust2', 'd8': 'Dust2', 'd9': 'Dust2',
    'd10': 'Dust2', 'd11': 'Dust2', 'd12': 'Dust2', 'd13': 'Dust2',
    'd14': 'Dust2', 'd15': 'Dust2', 'd16': 'Dust2', 'd17': 'Dust2',
    'd18': 'Dust2', 'd19': 'Dust2', 'd20': 'Dust2', 'd21': 'Dust2',
    'd22': 'Dust2', 'd23': 'Dust2', 'd24': 'Dust2', 'd25': 'Dust2',
    'anb': 'Anubis', 'nuke': 'Nuke',
    'mrg': 'Mirage', 'train': 'Train', 'inferno': 'Inferno',
    'dust2': 'Dust2', 'mirage': 'Mirage', 'anubis': 'Anubis',
}

ROUND_TYPES = {
    'P': '手枪局', 'pistol': '手枪局',
    'F': '长枪局', 'full-buy': '长枪局',
    'A': '强钢局', 'anti-eco': '强钢局',
    'H': '半起局', 'half-buy': '半起局',
    'E': '纯ECO', 'eco': '纯ECO',
}

SIDE_MAP = {'CT': 'CT', 'T': 'T', 'ct': 'CT', 't': 'T'}

# Issue classification keywords
ISSUE_PATTERNS = [
    (['丢呲', '没爆开', '白队友', '白自己', '烟雾呲', '闪丢', '火呲', '炸门',
      '烟雾弹', '烟墙', '灭烟', '灭火烟', '燃烧弹', '火瓶', '烧火', '火焰',
      '闪光弹', '闪光', '闪白', '全白', '白到',
      '手雷', '高爆', '炸雷', '丢错', '没丢好', '没丢到', '没丢中', '没丢出', '没丢进',
      '扔呲', '丢冒烟', '道具失误', '道具丢', '道具管', '烟丢', '火丢', '闪丢', '雷丢'], 'grenade'),
    (['走位', '撞一起', '漏点', '没观察', '站位', 'peek', '不该', '没到位',
      '蹲', '蹭', '架', '让', '回防', '拉', '压', '顶', '前压', '卡', '穿', '碰', '撞',
      '没看过', '没看', '选位'], 'position'),
    (['枪法', '没对过', '没打过', '打不过', '准', '瞄', '马枪', '空枪',
      '对枪', '拼枪', '没杀掉', '没打死', '没对', '对不过'], 'aim'),
    (['沟通', '信息', '报', '没说', '交流', '喊', 'call', '指挥', '没叫',
      '慌张', '交流不好', '没交流'], 'comms'),
    (['忘战术', '没发布', '未执行', '脱节', '没配合', '配合', '协同',
      '决策', '思路', '判断', '选错', '没听', '纪律', '战术'], 'tactics'),
]

SEVERITY_UPGRADE = ['全被', '全部', '忘', '未执行', '没发布', '违反', '直接违反']

# ============================================================
# DATABASE HELPERS
# ============================================================
conn = sqlite3.connect(DB_PATH)
conn.execute("PRAGMA foreign_keys = ON")
conn.execute("PRAGMA journal_mode = WAL")

# Only consider valid active players + staff (filter out junk from match imports)
VALID_PLAYERS = {'doomer', 'drace', '0z', 'glong', '4ever', 'hz', 'HZ'}

# Load player mapping from database
_rows = conn.execute(
    "SELECT nickname, real_name FROM players WHERE division='cs2'"
).fetchall()
for _row in _rows:
    PLAYER_MAP[_row[0]] = _row[1]
    PLAYER_MAP[_row[0].lower()] = _row[1]  # case-insensitive lookup
print(f"[ETL] Loaded {len(_rows)} players from database")


def execute(sql, params=()):
    return conn.execute(sql, params)

# Data integrity blocklists — skip these placeholder/header values
BLOCKED_OPPONENTS = {'OPPONENT', '未知', '___', 'match_data', 'MATCH_DATA'}
BLOCKED_MAP_NAMES = {'', 'UR 队员统计', '地图'}

# Opponent name normalization — unify common misspellings and aliases
OPPONENT_ALIASES = {
    'mongolza': 'Mongolz.A',
    'mongolz academy': 'Mongolz.A',
    'mongolz.academy': 'Mongolz.A',
    'nexvoid': 'NEXTVOID',
    'nextvoid': 'NEXTVOID',
    'thecube': 'THE QUBE',
    'the qube': 'THE QUBE',
    'the': 'THE QUBE',  # Truncated sheet name: "0608_vs_the"
}

def normalize_opponent(name):
    """Normalize opponent name using alias map (case-insensitive)."""
    if not name:
        return name
    key = name.strip().lower()
    return OPPONENT_ALIASES.get(key, name.strip())


def get_or_create_session(match_date, opponent, event_name=None):
    """Get or create a training session. Skip placeholder opponents."""
    if not opponent or opponent.strip() in BLOCKED_OPPONENTS or not match_date or match_date.strip() == '':
        print(f"[ETL] 跳过无效session: date={match_date}, opponent={opponent}")
        return None
    opponent = normalize_opponent(opponent)
    row = execute(
        "SELECT id FROM training_sessions WHERE match_date = ? AND opponent = ?",
        (match_date, opponent)
    ).fetchone()
    if row:
        return row[0]
    cur = execute(
        "INSERT INTO training_sessions (match_date, opponent, event_name) VALUES (?, ?, ?)",
        (match_date, opponent, event_name)
    )
    conn.commit()
    return cur.lastrowid


results = {
    'tactics': 0,
    'briefing_items': 0,
    'sessions': [],
    'training_rounds': 0,
    'matches': 0,
    'player_stats': 0,
    'errors': [],
}

# ============================================================
# PART 1: Parse Tactics Master
# ============================================================
print("[1/4] Parsing Tactics Master...")

wb = load_workbook(TACTICS, read_only=True, data_only=True)
ws = wb['战术总表']

execute("DELETE FROM tactics")  # Full refresh

tactic_count = 0
for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
    if not row or len(row) < 7:
        continue  # Skip incomplete rows
    tactic_id = str(row[0] or '').strip()
    map_name = str(row[1] or '').strip()
    side = str(row[2] or '').strip()
    round_type = str(row[3] or '').strip()
    name = str(row[4] or '').strip()
    target = str(row[5] or '').strip()
    note = str(row[6] or '').strip()

    if not tactic_id or tactic_id == 'NaN':
        continue
    if side not in ('T', 'CT'):
        continue  # Skip rows with invalid side

    # Map abbreviations
    full_map = MAP_ALIASES.get(map_name.lower(), map_name)
    full_round_type = ROUND_TYPES.get(round_type, round_type)

    execute("""
        INSERT INTO tactics (tactic_id, map_name, team_side, round_type, category, name, description, details, version)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1)
    """, (tactic_id, full_map, side, full_round_type, round_type, name, target, note))
    tactic_count += 1

conn.commit()
results['tactics'] = tactic_count
print(f"  ✓ {tactic_count} tactics imported")


# ============================================================
# PART 2: Parse Briefing (赛讯简报)
# ============================================================
print("[2/4] Parsing Briefing...")

wb2 = load_workbook(BRIEFING, read_only=True, data_only=True)

briefing_count = 0

for sheet_name in wb2.sheetnames:
    ws2 = wb2[sheet_name]
    rows = list(ws2.iter_rows(values_only=True))
    if len(rows) < 4:
        continue

    # Parse opponent info from row 1
    info_text = str(rows[1][0] or '')
    opp_match = re.search(r'对手[：:]\s*([^(\n\r|]+)', info_text)
    opponent = opp_match.group(1).strip() if opp_match else '未知'
    opponent = normalize_opponent(opponent)

    # Build match date
    date_str = f"2026-{sheet_name[:2]}-{sheet_name[2:]}"

    # Get or create session
    session_id = get_or_create_session(date_str, opponent)
    if session_id is None:
        continue  # Skip blocked/invalid sessions
    results['sessions'].append({'id': session_id, 'date': date_str, 'opponent': opponent})

    # Clear old briefing items
    execute("DELETE FROM briefing_items WHERE session_id = ?", (session_id,))

    for i in range(3, len(rows)):
        r = rows[i]
        if not r or len(r) < 1: continue
        rtype = str(r[0] or '').strip()
        command = str(r[1] or '').strip() if len(r) > 1 else ''
        note = str(r[2] or '').strip() if len(r) > 2 else ''

        if not rtype:
            continue

        # Determine round_type from note (P/F/A/H/E)
        round_type_code = note.strip() if note in ('P', 'F', 'A', 'H', 'E', '硬性') else None
        full_round_type = ROUND_TYPES.get(round_type_code, round_type_code) if round_type_code and round_type_code in ROUND_TYPES else rtype

        priority_map = {'P': '核心', 'A': '重点', 'E': '一般', 'F': '一般', 'H': '一般', '硬性': '核心'}
        priority = priority_map.get(note.strip(), '一般')

        # Determine map from round type text
        map_name = ''
        side = 'CT'
        if 'M1' in rtype:
            if 'CT' in rtype:
                side = 'CT'
                # M1 is my pick, M2 is opponent pick
                # Check opponent info for map
                my_map_match = re.search(r'我方[：:]?\s*(\w+)', info_text)
                if my_map_match:
                    map_raw = my_map_match.group(1).strip().rstrip('/')
                    map_name = MAP_ALIASES.get(map_raw.lower(), map_raw)
            elif 'T' in rtype:
                side = 'T'
        elif 'M2' in rtype:
            if 'CT' in rtype:
                side = 'CT'
            elif 'T' in rtype:
                side = 'T'
            opp_map_match = re.search(r'对方[：:]?\s*(\w+)', info_text)
            if opp_map_match:
                map_raw = opp_map_match.group(1).strip().rstrip('/')
                map_name = MAP_ALIASES.get(map_raw.lower(), map_raw)

        execute("""
            INSERT INTO briefing_items
            (session_id, map_name, team_side, tactic_id, round_type, priority, instruction, notes, sort_order)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (session_id, map_name, side, None, full_round_type, priority, command, note, i - 3))
        briefing_count += 1

conn.commit()
results['briefing_items'] = briefing_count
print(f"  ✓ {briefing_count} briefing items from {len(wb2.sheetnames)} sheets")

# ============================================================
# PART 3: Parse Training Log
# ============================================================
print("[3/4] Parsing Training Log...")

wb3 = load_workbook(TRAINING_LOG, read_only=True, data_only=True)

def classify_issues(text):
    """Classify issues from raw text description."""
    if not text:
        return {'grenade': False, 'position': False, 'aim': False, 'comms': False, 'tactics': False}
    text = text.lower()
    result = {}
    for keywords, issue_type in ISSUE_PATTERNS:
        result[issue_type] = any(kw.lower() in text for kw in keywords)
    # 中闪死/闪死 不计入道具失误（是被闪死，不是自己丢道具失误）
    if ('闪死' in text or '中闪' in text) and result.get('grenade'):
        result['grenade'] = False
    return result

def determine_severity(text):
    """Determine issue severity."""
    if not text:
        return '一般'
    for kw in SEVERITY_UPGRADE:
        if kw in text:
            return '严重'
    return '一般'

def guess_side(map_name, round_num_str, command_text, igl_text, fd_text):
    """Guess team side from context."""
    text = (command_text + ' ' + igl_text + ' ' + fd_text).lower()
    ct_score = sum(1 for kw in ['防守', 'ct', '回防', '架', '守', '卡', '蹲守', '赌'] if kw in text)
    t_score = sum(1 for kw in ['进攻', 'rush', '爆', '进', '攻', '突破', '冲', '出'] if kw in text)
    
    # Also check map/round context
    # Default: first round is typically T for the team that starts as T
    if ct_score > t_score:
        return 'CT'
    if t_score > ct_score:
        return 'T'
    return 'T'  # Default

training_round_count = 0

for sheet_name in wb3.sheetnames:
    if sheet_name == 'Sheet2' or not sheet_name.strip():
        continue
    ws3 = wb3[sheet_name]
    rows = list(ws3.iter_rows(values_only=True))
    if len(rows) < 4:
        continue

    # Parse date and opponent from sheet name: 0526_vs_tyloo
    parts = re.match(r'^(\d{4})_vs_(.+)$', sheet_name, re.IGNORECASE)
    if not parts:
        continue
    date_key = parts.group(1)
    opponent = parts.group(2).replace('_', ' ')
    date_str = f"2026-{date_key[:2]}-{date_key[2:]}"

    session_id = get_or_create_session(date_str, opponent)
    if not any(s['id'] == session_id for s in results['sessions']):
        results['sessions'].append({'id': session_id, 'date': date_str, 'opponent': opponent})

    # Clear old rounds
    execute("DELETE FROM training_rounds WHERE session_id = ?", (session_id,))

    # Determine if old format (3 cols) or new format (9 cols)
    header_row = rows[2]
    num_cols = len([c for c in header_row if c is not None])

    for i in range(3, len(rows)):
        r = rows[i]
        if not r or len(r) < 2: continue
        map_raw = str(r[0] or '').strip().lower() if r[0] else ''
        round_id = str(r[1] or '').strip() if r[1] else ''
        if not map_raw or not round_id:
            continue

        map_name = MAP_ALIASES.get(map_raw, map_raw.title())

        # ===== 新格式 12 列 (A-L): 地图/回合/阵营/战术/执行度/教练点评/责任人/IGL指令/首死ID/首死时间/首死原因/胜负 =====
        team_side = str(r[2] or '').strip() if len(r) > 2 and r[2] else ''       # C: 阵营
        tactic = str(r[3] or '').strip() if len(r) > 3 and r[3] else ''            # D: 战术
        execution = str(r[4] or '').strip() if len(r) > 4 and r[4] else ''         # E: 执行度
        coach_note = str(r[5] or '').strip() if len(r) > 5 and r[5] else ''        # F: 教练点评
        responsible = str(r[6] or '').strip() if len(r) > 6 and r[6] else ''       # G: 责任人
        igl_text = str(r[7] or '').strip() if len(r) > 7 and r[7] else ''          # H: IGL指令
        fd_id = str(r[8] or '').strip() if len(r) > 8 and r[8] else ''             # I: 首死ID
        fd_time_raw = r[9] if len(r) > 9 else None                  # J: 首死时间
        fd_cause = str(r[10] or '').strip() if len(r) > 10 and r[10] else ''  # K: 首死原因
        round_result_raw = str(r[11] or '').strip() if len(r) > 11 and r[11] else ''  # L: 胜负

        # Parse time
        fd_time = ''
        fd_time_seconds = None
        if fd_time_raw:
            if hasattr(fd_time_raw, 'strftime'):
                total_seconds = fd_time_raw.hour * 60 + fd_time_raw.minute
                fd_min = total_seconds // 60
                fd_sec = total_seconds % 60
                fd_time = f"{fd_min}:{fd_sec:02d}"
                fd_time_seconds = total_seconds
            else:
                fd_time = str(fd_time_raw).strip()
                tm = re.match(r'^(\d+):(\d+)$', fd_time)
                if tm:
                    fd_time_seconds = int(tm.group(1)) * 60 + int(tm.group(2))
                    fd_time = f"{int(tm.group(1))}:{tm.group(2).zfill(2)}"

        # Build command_text from IGL + tactic + coach note
        cmd_parts = []
        if igl_text:
            cmd_parts.append(igl_text)
        if tactic:
            cmd_parts.append(f"战术: {tactic}")
        if coach_note:
            cmd_parts.append(coach_note)
        cmd_text = ' | '.join(cmd_parts) if cmd_parts else ''
        full_desc = cmd_text

        # Only keep first_death if time is 1:00~1:55 (60-115s into round)
        if fd_time_seconds is not None and 60 <= fd_time_seconds <= 115:
            first_death = f"{fd_id} @ {fd_time} {fd_cause}".strip() if (fd_id or fd_cause) else ''
        elif fd_time_seconds is not None:
            first_death = ''
        else:
            first_death = f"{fd_id} @ {fd_time} {fd_cause}".strip() if (fd_id or fd_cause) else ''

        # Side
        side_raw = team_side.upper()
        if side_raw in ('T', 'CT'):
            side = side_raw
        else:
            side = guess_side(map_name, round_id, cmd_text, igl_text, fd_cause)

        # Issues, players, round result - same logic as parseTrainingV3
        issue_text = (coach_note + ' ' + fd_cause + ' ' + igl_text + ' ' + (responsible or '')).lower()
        is_flash_kill = False
        for kw in ['闪死', '中闪', '被闪']:
            if kw in issue_text:
                is_flash_kill = True
                break
        issues = classify_issues(issue_text)
        if is_flash_kill:
            issues['grenade'] = False

        severity = determine_severity(issue_text)

        # Round result
        round_result = None
        if round_result_raw in ('胜', 'win'):
            round_result = 'win'
        elif round_result_raw in ('负', 'loss'):
            round_result = 'loss'

        # Players involved: from responsible column OR text inference
        players_involved = set()
        if responsible:
            # Handle delimiters: . / , space
            raw_parts = re.split(r'[./,\s]+', responsible)
            for pid in raw_parts:
                pid = pid.strip().lower()
                if pid and pid in [p.lower() for p in PLAYER_MAP]:
                    players_involved.add(pid)
        if not players_involved:
            # Infer from text (word-boundary match to avoid substring false positives)
            search_text = (fd_id + ' ' + igl_text + ' ' + coach_note).lower()
            for pid in PLAYER_MAP:
                if re.search(r'\b' + re.escape(pid.lower()) + r'\b', search_text):
                    players_involved.add(pid.lower())
        players_str = ','.join(sorted(players_involved)) if players_involved else None

        execute("""
            INSERT INTO training_rounds
            (session_id, round_number, map_name, team_side, round_type,
             command_text, execution_text, first_death_reason,
             issue_grenade, issue_position, issue_aim, issue_comms, issue_tactics,
             round_result, players_involved, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            session_id, round_id, map_name, side, None,
            cmd_text if cmd_text else full_desc,
            execution or None,
            first_death or None,
            1 if issues['grenade'] else 0,
            1 if issues['position'] else 0,
            1 if issues['aim'] else 0,
            1 if issues['comms'] else 0,
            1 if issues['tactics'] else 0,
            round_result,
            players_str,
            severity,
        ))

        training_round_count += 1

conn.commit()
results['training_rounds'] = training_round_count
print(f"  ✓ {training_round_count} training rounds across {len(wb3.sheetnames) - 1} sheets")


# ============================================================
# PART 4: Parse Match Data (CS2_Match_Data_May2026.xlsx)
# ============================================================
print("[4/4] Parsing Match Data...")

wb4 = load_workbook(MATCH_DATA, read_only=True, data_only=True)

match_count = 0
stat_count = 0

for sheet_name in wb4.sheetnames:
    if not sheet_name.strip():
        continue
    ws4 = wb4[sheet_name]
    rows = list(ws4.iter_rows(values_only=True))
    if len(rows) < 12:
        continue

    # Row 1: date
    date_raw = rows[1][1] if rows[1] and len(rows[1]) > 1 else None
    if hasattr(date_raw, 'strftime'):
        date_str = date_raw.strftime('%Y-%m-%d')
    else:
        date_str = str(date_raw or '').strip()

    # Row 2: map
    map_raw = str(rows[2][1] or '').strip() if rows[2] and len(rows[2]) > 1 else ''
    map_name = MAP_ALIASES.get(map_raw.lower(), map_raw)

    # Row 3: match ID
    match_id_raw = rows[3][1] if rows[3] and len(rows[3]) > 1 else None

    # Row 4: time
    time_raw = str(rows[4][1] or '').strip() if rows[4] and len(rows[4]) > 1 else ''

    # Row 5: UR side/position
    # Row 6: UR score
    our_score = rows[6][1] if rows[6] and len(rows[6]) > 1 else 0
    # Row 7: Opponent score
    their_score = rows[7][1] if rows[7] and len(rows[7]) > 1 else 0
    # Row 8: Result
    result_raw = str(rows[8][1] or '').strip() if rows[8] and len(rows[8]) > 1 else ''

    # Parse opponent from sheet name: 0525_TyLoo_M1 → TyLoo
    opp_match = re.search(r'^\d{4}_(.+?)_M\d$', sheet_name, re.IGNORECASE)
    opponent = opp_match.group(1).replace('_', ' ') if opp_match else sheet_name
    opponent = normalize_opponent(opponent)

    # Data integrity: skip placeholder/junk opponents and dates
    if not date_str or date_str.strip() == '':
        print(f"[ETL] 跳过 {sheet_name}: 无效日期")
        continue
    if opponent.strip() in BLOCKED_OPPONENTS:
        print(f"[ETL] 跳过 {sheet_name}: 无效对手名 \"{opponent}\"")
        continue
    if not map_name or map_name.strip() in BLOCKED_MAP_NAMES:
        print(f"[ETL] 跳过 {sheet_name}: 无效地图名 \"{map_name}\"")
        continue

    try:
        our_score = int(our_score) if our_score else 0
        their_score = int(their_score) if their_score else 0
    except (ValueError, TypeError):
        our_score, their_score = 0, 0

    # Get or create match in matches table
    # Check if match exists by date+opponent+map
    existing = execute(
        "SELECT id FROM matches WHERE match_date = ? AND opponent = ? AND map_name = ? AND match_type = 'scrim'",
        (date_str, opponent, map_name)
    ).fetchone()

    if existing:
        match_db_id = existing[0]
        # Update scores (result is auto-generated from scores)
        execute(
            "UPDATE matches SET our_score = ?, their_score = ? WHERE id = ?",
            (our_score, their_score, match_db_id)
        )
    else:
        cur = execute("""
            INSERT INTO matches (match_date, opponent, event_name, match_type, map_name, our_score, their_score, bo_format, division)
            VALUES (?, ?, ?, 'scrim', ?, ?, ?, NULL, 'cs2')
        """, (date_str, opponent, None, map_name, our_score, their_score))
        match_db_id = cur.lastrowid
    match_count += 1

    # Parse player stats (UR: rows 12-16, Opponent: rows 21-24)
    def parse_player_stats(start_row, team_prefix, stat_count_ref):
        for row_idx in range(start_row, min(start_row + 6, len(rows))):
            r = rows[row_idx]
            if not r or len(r) < 5:
                continue
            player_name = str(r[1] or '').strip() if len(r) > 1 and r[1] else ''
            if not player_name or player_name == 'NaN' or player_name in ('玩家ID', '#', '合计'):
                continue
            kills = int(r[2]) if len(r) > 2 and r[2] and str(r[2]) != 'NaN' else 0
            deaths = int(r[3]) if len(r) > 3 and r[3] and str(r[3]) != 'NaN' else 0
            assists = int(r[4]) if len(r) > 4 and r[4] and str(r[4]) != 'NaN' else 0
            damage = int(r[5]) if len(r) > 5 and r[5] and str(r[5]) != 'NaN' else 0
            adr = float(r[6]) if len(r) > 6 and r[6] and str(r[6]) != 'NaN' else 0
            kd_ratio = float(r[7]) if len(r) > 7 and r[7] and str(r[7]) != 'NaN' else 0

            # Find player (do not auto-create; only match existing players)
            player_row = execute(
                "SELECT id FROM players WHERE nickname = ? AND division = 'cs2'",
                (player_name,)
            ).fetchone()
            if not player_row:
                print(f"[ETL] 跳过未知选手: {player_name}（不在players表中）")
                continue
            player_id = player_row[0]

            # Insert/update player stats
            existing_stat = execute(
                "SELECT id FROM player_stats WHERE match_id = ? AND player_id = ?",
                (match_db_id, player_id)
            ).fetchone()
            if existing_stat:
                execute(
                    "UPDATE player_stats SET kills = ?, deaths = ?, adr = ?, rating = ? WHERE id = ?",
                    (kills, deaths, adr, kd_ratio, existing_stat[0])
                )
            else:
                execute(
                    "INSERT INTO player_stats (match_id, player_id, kills, deaths, adr, rating) VALUES (?, ?, ?, ?, ?, ?)",
                    (match_db_id, player_id, kills, deaths, adr, kd_ratio)
                )
                stat_count_ref[0] += 1

    stat_count_list = [stat_count]
    # UR players start at row 12 (0-indexed)
    parse_player_stats(12, 'UR', stat_count_list)

    # Find opponent stats section (starts after "对手队员统计" label)
    opp_start = None
    for row_idx in range(18, min(len(rows), 26)):
        if rows[row_idx] and rows[row_idx][0] and '对手' in str(rows[row_idx][0]):
            opp_start = row_idx + 1
            break
    if opp_start:
        parse_player_stats(opp_start, 'OPP', stat_count_list)
    stat_count = stat_count_list[0]

conn.commit()
results['matches'] = match_count
results['player_stats'] = stat_count
print(f"  ✓ {match_count} matches, {stat_count} player stats")

# ============================================================
# Summary
# ============================================================
conn.close()

print("\n" + "=" * 60)
print("ETL COMPLETE")
print("=" * 60)
print(f"  战术总表: {results['tactics']} tactics")
print(f"  赛讯简报: {results['briefing_items']} briefing items")
print(f"  训练日志: {results['training_rounds']} rounds")
print(f"  比赛数据: {results['matches']} matches, {results['player_stats']} player stats")
print(f"  赛训赛次: {len(results['sessions'])} sessions")
if results['errors']:
    print(f"  ⚠ Errors: {results['errors']}")

# Output JSON for Node.js backend
print("\n__JSON__START__")
print(json.dumps(results, ensure_ascii=False, default=str))
print("__JSON__END__")
