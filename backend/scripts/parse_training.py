#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
parse_training.py - 解析训练日志 xlsx
  实际格式: 3列表格
  A=地图, B=回合(如R1), C=回合指令/执行记录(大白话混合)
输入: xlsx 文件路径 (sys.argv[1])
输出: JSON { "match_date": "...", "opponent": "...", "rounds": [...] }
"""
import sys, json, re, os
from openpyxl import load_workbook

# 地图名中文→英文
MAP_CN = {
    '荒漠迷城': 'Mirage', '米垃圾': 'Mirage',
    '远古遗迹': 'Ancient', '遗迹': 'Ancient',
    '核子危机': 'Nuke', 'nuke': 'Nuke',
    '阿努比斯': 'Anubis',
    '炙热沙城': 'Dust2', '沙二': 'Dust2',
    '炼狱小镇': 'Inferno', '小镇': 'Inferno',
    '死亡游乐园': 'Overpass', '游乐园': 'Overpass',
    '列车停放站': 'Train', '列车': 'Train',
    '殒命大厦': 'Vertigo', '大厦': 'Vertigo',
}

# 英文地图名标准化
MAP_EN = {m.lower(): m for m in
    ['Ancient', 'Anubis', 'Inferno', 'Mirage', 'Nuke',
     'Dust2', 'Train', 'Overpass', 'Vertigo', 'Cache']}

def normalize_map(name):
    if not name:
        return ''
    n = name.strip().lower()
    if n in MAP_EN:
        return MAP_EN[n]
    for cn, en in MAP_CN.items():
        if cn in name:
            return en
    # 尝试部分匹配
    for en_key, en_val in MAP_EN.items():
        if en_key in n:
            return en_val
    return name.strip()

# 关键词→问题类型
ISSUE_KW = {
    'grenade': ['道具', '丢呲', '白队友', '闪队友', '烟雾', '雷', '火', '闪光',
                '没丢', '丢的'],
    'position': ['走位', '站位', '位置', 'peek', '拉枪', 'timing', '没看',
                 '漏人', '没人看', '不看', '观察'],
    'aim': ['枪法', '对枪', '打不死', '空枪', 'miss', '瞄', '压枪', '没打过',
            '对不过', '没打赢', '击杀'],
    'comms': ['交流', '沟通', '报点', '信息', '没说', '没喊', '不交流',
              '不沟通', '不说话'],
    'tactics': ['战术', '配合', '执行', '没按', '没照', '脱离', '脱节', '没接',
                '没跟上', '没配合', '脱节', '不协调', '各打各'],
}

# 局型推断关键词
ROUND_TYPE_KW = {
    '手枪': '手枪', 'pistol': '手枪',
    '半起': '半起', 'half': '半起',
    'ECO': 'ECO', 'eco': 'ECO',
    '强起': '强起', 'force': '强起',
    '长枪': '长枪', 'full': '长枪',
}

# 阵营推断关键词
SIDE_KW_T = ['进攻', 'T ', 'T方', '匪', 'rush', '打', '爆', '夹', '清', '抢']
SIDE_KW_CT = ['防守', 'CT ', 'CT方', '警', '守', '防', '回防', '架', '反清']


def detect_issues(text):
    if not text:
        return {}
    issues = {}
    for key, keywords in ISSUE_KW.items():
        for kw in keywords:
            if kw in text:
                issues[key] = True
                break
    return issues


def infer_round_type(text):
    if not text:
        return ''
    t = text.lower()
    for kw, rt in ROUND_TYPE_KW.items():
        if kw.lower() in t:
            return rt
    return ''


def infer_side(text):
    if not text:
        return 'CT'
    ct_score = sum(1 for kw in SIDE_KW_CT if kw in text)
    t_score = sum(1 for kw in SIDE_KW_T if kw in text)
    if t_score > ct_score:
        return 'T'
    return 'CT'


def infer_result(text):
    if not text:
        return None
    t = text.lower()
    win_kw = ['win', '胜', '赢', 'ok', '成功', '拿下', '打赢', '吃掉']
    loss_kw = ['loss', '负', '输', '败', 'fail', '丢', '没赢', '被']
    for kw in win_kw:
        if kw in t:
            return 'win'
    for kw in loss_kw:
        if kw in t:
            return 'loss'
    return None


def parse_training_log(filepath):
    wb = load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    rounds = []
    match_date = None
    opponent = None
    current_map = ''

    # 文件名提取日期/对手
    fname = os.path.basename(filepath)
    m = re.search(r'(\d{4})', fname)
    if m:
        d = m.group(1)
        try:
            match_date = f"2026-{int(d[0:2]):02d}-{int(d[2:4]):02d}"
        except:
            pass
    m = re.search(r'(?:vs[._]?|VS[._]?)([A-Za-z]+)', fname)
    if m:
        opponent = m.group(1).capitalize()

    # 跳过前3行（标题+说明+表头）
    try:
        next(rows_iter)  # row 1: 标题
        next(rows_iter)  # row 2: 填写说明
        next(rows_iter)  # row 3: 表头
    except StopIteration:
        pass

    for row in rows_iter:
        if not row or not any(row):
            continue

        values = [str(c).strip() if c is not None else '' for c in row]

        # Col A: 地图
        map_raw = values[0] if len(values) > 0 else ''
        # Col B: 回合编号 (R1, R2...)
        round_label = values[1] if len(values) > 1 else ''
        # Col C: 指令/执行记录
        text = values[2] if len(values) > 2 else ''

        if not round_label:
            # 可能是地图分隔行
            if map_raw:
                normalized = normalize_map(map_raw)
                if normalized:
                    current_map = normalized
            continue

        # 解析回合编号
        rn_match = re.match(r'R\s*(\d+)', round_label, re.IGNORECASE)
        if not rn_match:
            # 也可能是纯数字
            rn_match = re.match(r'(\d+)', round_label)
        if not rn_match:
            continue

        round_number = int(rn_match.group(1))

        # 地图名
        if map_raw and map_raw.lower() in MAP_EN:
            current_map = MAP_EN[map_raw.lower()]
        map_name = current_map or normalize_map(map_raw)

        # 从文本推断属性
        team_side = infer_side(text) if text else 'CT'
        round_type = infer_round_type(text)
        result = infer_result(text)
        detected = detect_issues(text)

        # 尝试从文本中分离"首死"信息
        first_death = None
        sds = re.search(r'(?:首死|率先阵亡)[：:]*\s*(.+?)(?:[。，,\.]|$)', text)
        if sds:
            first_death = sds.group(1).strip()

        rounds.append({
            'round_number': round_number,
            'map_name': map_name,
            'team_side': team_side,
            'round_type': round_type,
            'command_text': text,
            'execution_text': '',
            'first_death_reason': first_death,
            'issue_grenade': detected.get('grenade', False),
            'issue_position': detected.get('position', False),
            'issue_aim': detected.get('aim', False),
            'issue_comms': detected.get('comms', False),
            'issue_tactics': detected.get('tactics', False),
            'round_result': result,
            'notes': None,
        })

    wb.close()

    if not rounds:
        return {'error': '未解析到回合数据，请检查格式（需 A=地图 B=回合 C=指令）'}

    stats = {
        'total': len(rounds),
        'issues': sum(1 for r in rounds if any([
            r['issue_grenade'], r['issue_position'], r['issue_aim'],
            r['issue_comms'], r['issue_tactics']
        ])),
        'maps': list(set(r['map_name'] for r in rounds if r['map_name'])),
    }

    return {
        'match_date': match_date,
        'opponent': opponent or 'Unknown',
        'rounds': rounds,
        'stats': stats,
    }


if __name__ == '__main__':
    if len(sys.argv) < 2:
        json.dump({'error': '缺少文件路径参数'}, sys.stdout, ensure_ascii=False)
        sys.exit(1)
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass
    try:
        result = parse_training_log(sys.argv[1])
        json.dump(result, sys.stdout, ensure_ascii=False)
    except Exception as e:
        json.dump({'error': str(e)}, sys.stdout, ensure_ascii=False)
        sys.exit(1)
