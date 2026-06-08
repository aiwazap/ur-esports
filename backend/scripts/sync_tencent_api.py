"""
腾讯文档实时同步脚本 v2
通过 export API 导出在线表格为 XLSX → 下载 → 用现有 ETL 逻辑解析 → 写入 SQLite
"""
import json, sqlite3, re, os, sys, time, requests, io, tempfile
from openpyxl import load_workbook

API_URL = "https://docs.qq.com/openapi/mcp"
API_KEY = "95e588f104e6471d8c3c02f9401445c6"
FOLDER_ID = "RGiUuasADqZb"

MAP_ALIASES = {
    'overpass': 'Overpass', 'op': 'Overpass', 'ancient': 'Ancient', 'anc': 'Ancient',
    'd2': 'Dust2', 'nuke': 'Nuke', 'mrg': 'Mirage', 'train': 'Train',
    'inferno': 'Inferno', 'dust2': 'Dust2', 'mirage': 'Mirage', 'anubis': 'Anubis',
}
ROUND_TYPES = {'P': '手枪局', 'F': '长枪局', 'A': '强钢局', 'H': '半起局', 'E': '纯ECO'}
SIDE_MAP = {'CT': 'CT', 'T': 'T', 'ct': 'CT', 't': 'T'}
BLOCKED_OPPONENTS = {'OPPONENT', '未知', '___', 'match_data', 'MATCH_DATA'}
BLOCKED_MAPS = {'', 'UR 队员统计', '地图'}
OPPONENT_ALIASES = {
    'mongolza': 'Mongolz.A', 'mongolz academy': 'Mongolz.A', 'mongolz.academy': 'Mongolz.A',
    'nexvoid': 'NEXTVOID', 'nextvoid': 'NEXTVOID',
    'thecube': 'THE QUBE', 'the qube': 'THE QUBE', 'the': 'THE QUBE',
}
ISSUE_PATTERNS = [
    (['丢呲','没爆开','白队友','烟雾呲','闪丢','火呲','炸门','烟雾弹','烟墙','灭烟',
      '灭火烟','燃烧弹','火瓶','烧火','闪光弹','闪光','闪白','手雷','高爆','丢错',
      '没丢好','没丢到','没丢中','没丢出','没丢进','扔呲','丢冒烟','道具失误','道具丢','道具管',
      '烟丢','火丢','闪丢','雷丢','闪光失误'], 'grenade'),
    (['走位','撞一起','漏点','没观察','站位','peek','不该','没到位','蹲','蹭','架',
      '让','回防','拉','压','顶','前压','卡','穿','碰','撞','没看过','没看','选位'], 'position'),
    (['枪法','没对过','没打过','打不过','准','瞄','马枪','空枪','对枪','拼枪',
      '没杀掉','没打死','没对','对不过'], 'aim'),
    (['沟通','信息','报','没说','交流','喊','call','指挥','没叫','慌张','交流不好','没交流'], 'comms'),
    (['战术','没按','不按','违抗','擅自','策划','方案','指令','违反','未执行','没发布'], 'tactics'),
]
KNOWN_MAPS = {'Mirage','Dust2','Inferno','Nuke','Ancient','Anubis','Overpass','Vertigo','Train'}
VALID_PLAYERS = {'doomer','Doomer','drace','0z','glong','gLong','4ever','hz','HZ'}

def call_api(method, arguments):
    resp = requests.post(API_URL, json={
        'jsonrpc': '2.0', 'method': 'tools/call',
        'params': {'name': method, 'arguments': arguments},
        'id': 1
    }, headers={'Authorization': API_KEY, 'Content-Type': 'application/json'}, timeout=120)
    data = resp.json()
    text = data.get('result',{}).get('content',[{}])[0].get('text','{}')
    return json.loads(text)

def export_and_download(file_id, fmt='xlsx'):
    """Export a Tencent Docs spreadsheet, poll progress, download content."""
    task = call_api('manage.export_file', {'file_id': file_id, 'format': fmt})
    task_id = task.get('task_id')
    if not task_id:
        print(f"  ⚠ Export failed: {task}")
        return None, None
    
    for _ in range(30):
        time.sleep(2)
        progress = call_api('manage.export_progress', {'task_id': task_id})
        if progress.get('progress') == 100:
            url = progress.get('file_url', '')
            if url:
                resp = requests.get(url, timeout=120)
                if resp.status_code == 200:
                    print(f"  ✓ Downloaded {len(resp.content)} bytes ({fmt})")
                    return resp.content, fmt
            break
    
    print(f"  ⚠ Export timeout or no URL")
    return None, None

def parse_csv_sheets(csv_data):
    """Parse CSV bytes into dict of sheet_name -> [rows] (tab-separated sheets)"""
    # For CSV export, all sheets are concatenated or returned as one CSV
    # Actually Tencent Docs exports each sheet as a separate CSV file
    # When format=csv, we get the entire workbook... but multi-sheet workbooks don't export well
    # We'll handle this differently per file type
    text = csv_data.decode('utf-8', errors='replace')
    return text

def normalize_opponent(name):
    key = name.strip().lower()
    return OPPONENT_ALIASES.get(key, name.strip())

def classify_issues(text):
    if not text: return {}
    text_lower = text.lower()
    return {cat: True for patterns, cat in ISSUE_PATTERNS if any(pat in text_lower for pat in patterns)}

def sync_tactics(conn, xlsx_data):
    """Parse tactics from XLSX bytes"""
    print("[1/4] Syncing Tactics...")
    if not xlsx_data: return 0
    
    wb = safe_load_workbook(xlsx_data, 'briefing')
    if not wb: return 0
    ws = wb['战术总表'] if '战术总表' in wb.sheetnames else wb[wb.sheetnames[0]]
    
    conn.execute("DELETE FROM tactics")
    count = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 7: continue
        tid = str(row[0] or '').strip()
        if not tid or tid == 'NaN': continue
        side = str(row[2] or '').strip()
        if side not in ('T', 'CT'): continue
        
        map_name = MAP_ALIASES.get(str(row[1] or '').lower(), str(row[1] or ''))
        round_type = ROUND_TYPES.get(str(row[3] or '').strip(), str(row[3] or '').strip())
        name = str(row[4] or '').strip()
        target_area = str(row[5] or '').strip()
        note = str(row[6] or '').strip()
        
        conn.execute(
            "INSERT INTO tactics (tactic_id, map_name, team_side, round_type, category, name, description, details, version) VALUES (?,?,?,?,?,?,?,?,?)",
            (tid, map_name, side, round_type, '', name, target_area, note, 'v2'))
        count += 1
    wb.close()
    conn.commit()
    print(f"  ✓ {count} tactics")
    return count

def safe_load_workbook(xlsx_data, label=''):
    """Load XLSX, stripping broken style definitions if needed."""
    import zipfile, warnings, io
    warnings.filterwarnings('ignore')
    try:
        return load_workbook(io.BytesIO(xlsx_data), read_only=True, data_only=True, keep_links=False)
    except:
        pass
    # Fallback: strip problematic styles.xml from the ZIP
    try:
        buf = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(xlsx_data), 'r') as zin:
            with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    if 'styles.xml' in item.filename:
                        zout.writestr(item, '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"/>')
                    else:
                        zout.writestr(item, zin.read(item.filename))
        buf.seek(0)
        return load_workbook(buf, read_only=True, data_only=True, keep_links=False)
    except Exception as e:
        print(f"  ⚠ Cannot load {label}: {e}")
        return None

def sync_briefing(conn, xlsx_data):
    """Parse briefing from XLSX bytes"""
    print("[2/4] Syncing Briefing...")
    if not xlsx_data: return 0
    
    wb = safe_load_workbook(xlsx_data, 'briefing')
    if not wb: return 0
    count = 0
    sessions = 0
    
    for sheet_name in wb.sheetnames:
        if not re.match(r'^\d{4}$', sheet_name):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2: continue
        
        # Row 1: opponent info
        info = str(rows[1][0] or '') if len(rows) > 1 else ''
        opp_match = re.search(r'对手[：:]\s*([^(\n\r|]+)', info)
        opponent = opp_match.group(1).strip() if opp_match else '未知'
        opponent = normalize_opponent(opponent)
        date_str = f"2026-{sheet_name[:2]}-{sheet_name[2:]}"
        
        if opponent in BLOCKED_OPPONENTS:
            continue
        
        # Get or create session
        row = conn.execute("SELECT id FROM training_sessions WHERE match_date=? AND opponent=?", (date_str, opponent)).fetchone()
        if row:
            session_id = row[0]
        else:
            cur = conn.execute("INSERT INTO training_sessions (match_date, opponent) VALUES (?,?)", (date_str, opponent))
            session_id = cur.lastrowid
        
        conn.execute("DELETE FROM briefing_items WHERE session_id=?", (session_id,))
        
        # Parse items (start from row 3, skipping headers)
        for i in range(3, len(rows)):
            r = rows[i]
            if not r or len(r) < 1: continue
            rtype = str(r[0] or '').strip()
            if not rtype: continue
            cmd = str(r[1] or '').strip() if len(r) > 1 else ''
            note = str(r[2] or '').strip() if len(r) > 2 else ''
            
            round_type = ROUND_TYPES.get(rtype, rtype)
            tactic_id = None
            if cmd:
                tm = re.search(r'([A-Z]+_T_[FAHEP]\d+)', cmd, re.IGNORECASE)
                if tm: tactic_id = tm.group(1).upper()
            
            conn.execute(
                "INSERT INTO briefing_items (session_id, map_name, team_side, round_type, instruction, notes, tactic_id, sort_order) VALUES (?,?,'T',?,?,?,?,?)",
                (session_id, '', round_type, cmd, note, tactic_id, i - 2))
            count += 1
        sessions += 1
        print(f"  ✓ {date_str} {opponent}: {count} items")
    
    wb.close()
    conn.commit()
    print(f"  ✓ {count} briefing items, {sessions} sessions")
    return count

def sync_training_log(conn, xlsx_data):
    """Parse training log from XLSX bytes"""
    print("[3/4] Syncing Training Log...")
    if not xlsx_data: return 0
    
    wb = safe_load_workbook(xlsx_data, 'briefing')
    if not wb: return 0
    count = 0
    sessions = 0
    
    for sheet_name in wb.sheetnames:
        parts = re.match(r'^(\d{4})_vs_(.+)$', sheet_name, re.IGNORECASE)
        if not parts: continue
        
        date_key = parts.group(1)
        opponent_raw = parts.group(2).replace('_', ' ').strip()
        opponent = normalize_opponent(opponent_raw)
        date_str = f"2026-{date_key[:2]}-{date_key[2:]}"
        
        if opponent in BLOCKED_OPPONENTS: continue
        
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4: continue
        
        # Get or create session
        row = conn.execute("SELECT id FROM training_sessions WHERE match_date=? AND opponent=?", (date_str, opponent)).fetchone()
        if row:
            session_id = row[0]
        else:
            cur = conn.execute("INSERT INTO training_sessions (match_date, opponent) VALUES (?,?)", (date_str, opponent))
            session_id = cur.lastrowid
        
        conn.execute("DELETE FROM training_rounds WHERE session_id=?", (session_id,))
        
        # Parse rounds (start from row 3)
        for i in range(3, len(rows)):
            r = rows[i]
            if not r or len(r) < 2: continue
            map_raw = str(r[0] or '').strip().lower()
            round_id = str(r[1] or '').strip()
            if not map_raw or not round_id: continue
            
            map_name = MAP_ALIASES.get(map_raw, map_raw.title())
            team_side = SIDE_MAP.get(str(r[2] or '').strip(), str(r[2] or '').strip()) if len(r) > 2 else ''
            if team_side not in ('T', 'CT'):
                team_side = 'T'  # Default for CHECK constraint
            tactic = str(r[3] or '').strip() if len(r) > 3 else ''
            coach_note = str(r[5] or '').strip() if len(r) > 5 else ''
            responsible = str(r[6] or '').strip() if len(r) > 6 else ''
            fd_id = str(r[8] or '').strip() if len(r) > 8 else ''
            fd_time = str(r[9] or '').strip() if len(r) > 9 else ''
            fd_cause = str(r[10] or '').strip() if len(r) > 10 else ''
            round_result = str(r[11] or '').strip() if len(r) > 11 else ''
            if round_result.lower() not in ('win', 'loss'):
                round_result = 'loss'  # Default for CHECK constraint
            
            issues = classify_issues(f"{coach_note} {fd_cause}")
            # Build first_death_reason: fd_id @ fd_time fd_cause
            fd_cause_clean = (fd_id + ' @ ' + fd_time + ' ' + fd_cause).strip() if fd_id else ''
            players = set()
            for pid in VALID_PLAYERS:
                if pid.lower() in (responsible or '').lower():
                    players.add(pid)
            if not players and fd_id and fd_id.lower() in VALID_PLAYERS:
                players.add(fd_id)
            
            conn.execute("""INSERT INTO training_rounds 
                (session_id, round_number, map_name, team_side, 
                 command_text, first_death_reason, round_result,
                 players_involved, issue_grenade, issue_position, issue_aim, issue_comms, issue_tactics)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (session_id, round_id, map_name, team_side, 
                 coach_note, fd_cause_clean, round_result,
                 ','.join(players) if players else '',
                 1 if 'grenade' in issues else 0, 1 if 'position' in issues else 0,
                 1 if 'aim' in issues else 0, 1 if 'comms' in issues else 0,
                 1 if 'tactics' in issues else 0))
            count += 1
        
        sessions += 1
        print(f"  ✓ {date_str} {opponent}: {count} rounds")
    
    wb.close()
    conn.commit()
    print(f"  ✓ {count} training rounds, {sessions} sessions")
    return count

def sync_match_data(conn, xlsx_data):
    """Parse match data from XLSX bytes"""
    print("[4/4] Syncing Match Data...")
    if not xlsx_data: return 0
    
    wb = safe_load_workbook(xlsx_data, 'briefing')
    if not wb: return 0
    match_count = 0
    stat_count = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows: continue
        
        # Check format: first cell should be "CS2 比赛数据" or old format
        first_cell = str(rows[0][0] or '') if rows[0] else ''
        
        if first_cell.strip() == 'CS2 比赛数据':
            # New format
            date_raw = str(rows[1][1] or '').strip() if len(rows) > 1 and len(rows[1]) > 1 else ''
            date_str = date_raw.split(' ')[0] if date_raw else ''
            map_name = str(rows[2][1] or '').strip() if len(rows) > 2 and len(rows[2]) > 1 else ''
            
            opp_match = re.search(r'^\d{4}_(.+?)_M\d$', sheet_name, re.IGNORECASE)
            opponent = opp_match.group(1).replace('_', ' ') if opp_match else sheet_name
            opponent = normalize_opponent(opponent)
            
            if not date_str or date_str == '地图' or opponent in BLOCKED_OPPONENTS or not map_name or map_name in BLOCKED_MAPS:
                print(f"  ⚠ Skip {sheet_name}: invalid data")
                continue
            
            our_score = int(rows[6][1]) if len(rows) > 6 and len(rows[6]) > 1 and rows[6][1] else 0
            their_score = int(rows[7][1]) if len(rows) > 7 and len(rows[7]) > 1 and rows[7][1] else 0
            
            existing = conn.execute(
                "SELECT id FROM matches WHERE match_date=? AND opponent=? AND map_name=? AND match_type='scrim'",
                (date_str, opponent, map_name)).fetchone()
            if existing:
                conn.execute("UPDATE matches SET our_score=?, their_score=? WHERE id=?", (our_score, their_score, existing[0]))
                match_id = existing[0]
            else:
                cur = conn.execute(
                    "INSERT INTO matches (match_date, opponent, match_type, map_name, our_score, their_score, bo_format, division) VALUES (?,?,'scrim',?,?,?,NULL,'cs2')",
                    (date_str, opponent, map_name, our_score, their_score))
                match_id = cur.lastrowid
            match_count += 1
            
            # Player stats
            for ri in range(12, min(17, len(rows))):
                r = rows[ri]
                if not r or len(r) < 2: continue
                pname = str(r[1] or '').strip()
                if not pname or pname in ('NaN', '玩家ID', '#', '合计'): continue
                kills = int(r[2]) if len(r) > 2 and r[2] else 0
                deaths = int(r[3]) if len(r) > 3 and r[3] else 0
                adr = float(r[6]) if len(r) > 6 and r[6] else 0
                rating = float(r[7]) if len(r) > 7 and r[7] else 0
                
                pr = conn.execute("SELECT id FROM players WHERE nickname=? AND division='cs2'", (pname,)).fetchone()
                if not pr: continue
                
                es = conn.execute("SELECT id FROM player_stats WHERE match_id=? AND player_id=?", (match_id, pr[0])).fetchone()
                if es:
                    conn.execute("UPDATE player_stats SET kills=?,deaths=?,adr=?,rating=? WHERE id=?", (kills, deaths, adr, rating, es[0]))
                else:
                    conn.execute("INSERT INTO player_stats (match_id, player_id, kills, deaths, adr, rating) VALUES (?,?,?,?,?,?)",
                                (match_id, pr[0], kills, deaths, adr, rating))
                    stat_count += 1
            print(f"  ✓ {date_str} {opponent} {map_name}: {our_score}-{their_score}")
    
    wb.close()
    conn.commit()
    print(f"  ✓ {match_count} matches, {stat_count} player stats")
    return match_count

def main():
    db_path = os.environ.get('DB_PATH', os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'ur_esports.db'))
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA foreign_keys = ON")
    
    # Discover file IDs
    folder = call_api('manage.folder_list', {'folder_id': FOLDER_ID})
    file_ids = {}
    for f in folder.get('list', []):
        title = f.get('title', '').lower()
        if '训练日志' in title: file_ids['training'] = f['id']
        elif '每日简报' in title: file_ids['briefing'] = f['id']
        elif '战术总表' in title: file_ids['tactics'] = f['id']
        elif 'match_data' in title: file_ids['match_data'] = f['id']
    print(f"[SYNC] Files: {json.dumps(file_ids, ensure_ascii=False)}")
    
    results = {}
    
    # Export and sync each file
    for key, label in [('tactics', '战术总表'), ('briefing', '每日简报'), ('training', '训练日志'), ('match_data', '比赛数据')]:
        if key not in file_ids:
            print(f"[{key}] 未找到")
            continue
        print(f"\n=== {label} ===")
        xlsx, _ = export_and_download(file_ids[key])
        if xlsx:
            if key == 'tactics': results['tactics'] = sync_tactics(conn, xlsx)
            elif key == 'briefing': results['briefing'] = sync_briefing(conn, xlsx)
            elif key == 'training': results['rounds'] = sync_training_log(conn, xlsx)
            elif key == 'match_data': results['matches'] = sync_match_data(conn, xlsx)
        else:
            print(f"  ⚠ 导出失败")
    
    conn.close()
    print("\n__JSON__START__")
    print(json.dumps(results, ensure_ascii=False))
    print("__JSON__END__")

if __name__ == '__main__':
    main()
