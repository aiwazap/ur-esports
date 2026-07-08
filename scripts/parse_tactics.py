#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
parse_tactics.py - 解析战术总表 xlsx
  实际格式: Sheet1(战术总表) = 所有五图数据合并在一个Sheet
  表头在第3行, 数据从第4行开始
  A=战术编号, B=地图代码, C=阵营, D=局型代码, E=战术名, F=默认目标, G=备注
输入: xlsx 文件路径 (sys.argv[1])
输出: JSON { "tactics": [...] } 到 stdout
"""
import sys, json, re
from openpyxl import load_workbook

# 地图代码 → 标准地图名
MAP_ALIASES = {
    'mrg': 'Mirage', 'mir': 'Mirage', 'mirage': 'Mirage',
    'anc': 'Ancient', 'ancient': 'Ancient',
    'nuke': 'Nuke',
    'anb': 'Anubis', 'anubis': 'Anubis',
    'd2': 'Dust2', 'dust2': 'Dust2',
    'inf': 'Inferno', 'inferno': 'Inferno',
    'ovp': 'Overpass', 'overpass': 'Overpass',
    'trn': 'Train', 'train': 'Train',
    'vtg': 'Vertigo', 'vertigo': 'Vertigo',
}

# 局型代码 → 中文名
ROUND_TYPE_MAP = {
    'f': '长枪', 'g': '长枪',
    'h': '半起',
    'e': 'ECO',
    'p': '手枪',
    'a': '强起',
}

# 阵营映射 (中文 → 英文)
SIDE_MAP = {
    '进攻方': 'T', '进攻': 'T', 't': 'T',
    '防守方': 'CT', '防守': 'CT', 'ct': 'CT',
}

# 合法战术编号正则: 如 MRG-T-F01, ANC-CT-H02
TACTIC_ID_PATTERN = re.compile(
    r'^[A-Z0-9]{2,4}-(?:T|CT)-(?:F|H|E|P|A|G)\d{2}$', re.IGNORECASE
)


def parse_tactics(filepath):
    wb = load_workbook(filepath, data_only=True, read_only=True)
    tactics = []
    seen = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # 跳过元数据 Sheet（如"代表队名称及选手"）
        sn_lower = sheet_name.strip().lower()
        if any(kw in sn_lower for kw in ('选手', '代表队', '说明', '注释', 'player')):
            continue

        rows_iter = ws.iter_rows(min_row=1, values_only=True)

        # 跳过前两行（标题+说明），第3行是表头
        try:
            next(rows_iter)  # row 1: 标题
            next(rows_iter)  # row 2: 填写说明
            header = next(rows_iter)  # row 3: 表头
        except StopIteration:
            continue

        # 从表头确认列位置 (或使用固定位置: A=tactic_id, B=map, C=side, D=round_type, E=name, F=desc, G=notes)
        col_map = {0: 'tactic_id', 1: 'map_code', 2: 'side', 3: 'round_code',
                    4: 'name', 5: 'description', 6: 'notes'}
        for i, h in enumerate(header):
            if not h:
                continue
            h_str = str(h).strip()
            if '战术编号' in h_str or '编号' in h_str:
                col_map[i] = 'tactic_id'
            elif '地图' in h_str:
                col_map[i] = 'map_code'
            elif '角色' in h_str or '阵营' in h_str or '方' in h_str:
                col_map[i] = 'side'
            elif '局' in h_str or '类型' in h_str:
                col_map[i] = 'round_code'
            elif '战术名' in h_str or '名称' in h_str:
                col_map[i] = 'name'
            elif '目标' in h_str or '指令' in h_str or '描述' in h_str:
                col_map[i] = 'description'
            elif '备注' in h_str:
                col_map[i] = 'notes'

        # 收集各列索引
        idx = {}
        for i, role in col_map.items():
            idx[role] = i

        for row in rows_iter:
            if not row or not any(row):
                continue

            # 取各列值
            def cell_val(col_name):
                ci = idx.get(col_name, -1)
                if ci >= 0 and ci < len(row) and row[ci] is not None:
                    return str(row[ci]).strip()
                return ''

            tactic_id = cell_val('tactic_id')
            if not tactic_id or tactic_id == 'None':
                continue

            # 用正则验证是否为合法战术编号，过滤表头/说明行
            if not TACTIC_ID_PATTERN.match(tactic_id.upper()):
                continue

            tactic_id = tactic_id.upper()

            # 跳过重复
            if tactic_id in seen:
                continue
            seen.add(tactic_id)

            # 地图名映射
            map_code = cell_val('map_code').upper()
            map_name = MAP_ALIASES.get(map_code.lower(), map_code)

            # 阵营
            side_raw = cell_val('side').upper()
            team_side = SIDE_MAP.get(side_raw, side_raw if side_raw in ('T', 'CT') else 'CT')

            # 局型
            round_code = cell_val('round_code').lower()
            round_type = ROUND_TYPE_MAP.get(round_code, round_code.upper() if round_code else None)

            # 如果从列中没解析到局型，从编号中提取
            if not round_type:
                parts = tactic_id.split('-')
                if len(parts) >= 3:
                    rc = parts[2][0].lower()
                    round_type = ROUND_TYPE_MAP.get(rc)

            name = cell_val('name')
            description = cell_val('description')
            notes = cell_val('notes')

            # 分类 = 地图名 (原Excel中B列就是地图代码)
            category = map_name or ''

            tactics.append({
                'tactic_id': tactic_id,
                'map_name': map_name,
                'team_side': team_side,
                'round_type': round_type,
                'category': category,
                'name': name,
                'description': description,
                'notes': notes,
            })

    wb.close()

    if not tactics:
        return {'error': '未解析到任何战术数据，请检查文件格式'}

    return {
        'tactics': tactics,
        'count': len(tactics),
        'maps': list(set(t['map_name'] for t in tactics)),
    }


if __name__ == '__main__':
    if len(sys.argv) < 2:
        json.dump({'error': '缺少文件路径参数'}, sys.stdout, ensure_ascii=False)
        sys.exit(1)

    try:
        result = parse_tactics(sys.argv[1])
        # 显式设置 stdout 编码为 utf-8
        sys.stdout.reconfigure(encoding='utf-8')
        json.dump(result, sys.stdout, ensure_ascii=False)
    except Exception as e:
        try:
            sys.stdout.reconfigure(encoding='utf-8')
        except Exception:
            pass
        json.dump({'error': str(e)}, sys.stdout, ensure_ascii=False)
        sys.exit(1)
